#!/usr/bin/python3
#----------------------------------------------------------------------
#  File name     : excel.py
#  Function      : <empty>
#  Author        : Tian
#  Email         : tientq@coasia.com
#  Created time  : 2024-04-10 17:46:02
#----------------------------------------------------------------------
import re, os
import sys ; sys.path.append("/mnt/users/tientq/tiandir/python/")
import tian; sys.path.append(tian.user_library_path)
from   tian            import run, print_banner, excel
from   tian.global_var import *

# ----------------------------------------------------------------+
# Import                                                          |
# ----------------------------------------------------------------+
import xlrd
import openpyxl
from openpyxl                          import formatting, styles
from openpyxl.formatting.rule          import FormulaRule, CellIsRule
from openpyxl.styles                   import PatternFill, Font, NamedStyle
from openpyxl.styles.borders           import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers           import FORMAT_PERCENTAGE_00
from openpyxl.worksheet.table          import Table, TableStyleInfo
from openpyxl.worksheet.hyperlink      import Hyperlink
from .global_func                      import print_banner

#-------------------------------------------------------------------|
# Print color                                                       |
#-------------------------------------------------------------------|
BLACK   = '\033[30;1m'
RED     = '\033[31;1m'
GREEN   = '\033[32;1m'
YELLOW  = '\033[33;1m'
BLUE    = '\033[34;1m'
MAGENTA = '\033[35;1m'
CYAN    = '\033[36;1m'
WHITE   = '\033[37;1m'
RESET   = '\033[0m'

class excel:
    # ----------------------------------------------------------------|
    # Properties                                                      |
    # ----------------------------------------------------------------|
    font        = Font(color              = '227447')
    color_fill  = PatternFill(start_color = 'C6EFCE', end_color = 'C6EFCE', fill_type = 'solid')
    link        = ""
    sheet_names = []
    wb          = None
    table_idx   = 0

    # ----------------------------------------------------------------|
    # Methods                                                         |
    # ----------------------------------------------------------------|
    def __init__(self, excel_link):
        if ("xlsx" not in excel_link):
            print(f"Warning: Cannot open {excel_link}")
            print(f"Info: Cant not open xls format directly")
            print(f"      Please convert it to xlsx by: tian.excel.convert_xls_to_xlsx(xls_link)")
            exit()

        self.link = excel_link
        try:
            print_banner(f"{excel_link} is loading ...")
            self.wb          = openpyxl.load_workbook(f"{excel_link}")
            self.sheet_names = self.wb.get_sheet_names()
        except Exception as e:
            print(f"[open_xlsx_error]: {e}")
            os.system(f"echo {excel_link} >> open_excel_error.log")

    def get_sheet_by_name(self, sheet_name:str):
        try:
            return self.wb[sheet_name]
        except:
            print(f"{self.link}.{sheet_name} is not exist")

    def resize_col(self, cell, width:int):
        sheet = cell.parent
        sheet.column_dimensions[self.get_column_letter(cell)].width = width

    def resize_row(self, cell, height:int):
        sheet = cell.parent
        sheet.row_dimensions[cell.row].height = height

    def fmt_wrap_text(self, cell):
        cell.alignment = openpyxl.styles.Alignment(wrapText=True)
    
    def fmt_percent_style(self, cell):
        sheet              = cell.parent
        cell.number_format = '0.00%'
        self.get_color('red')
        rule_red           = CellIsRule(operator='between', formula=['0'  , '0.5'], stopIfTrue=True, fill=self.color_fill, font=self.font)
        self.get_color('yellow')
        rule_yellow        = CellIsRule(operator='between', formula=['0.5', '0.8'], stopIfTrue=True, fill=self.color_fill, font=self.font)
        self.get_color('green')
        rule_green         = CellIsRule(operator='between', formula=['0.8', '1'  ], stopIfTrue=True, fill=self.color_fill, font=self.font)
        sheet.conditional_formatting.add(cell.coordinate, rule_red)
        sheet.conditional_formatting.add(cell.coordinate, rule_yellow)
        sheet.conditional_formatting.add(cell.coordinate, rule_green)
    
    def drop_down_list(self, cell, choices:list):
        sheet            = cell.parent
        valid_options    = '"' + ','.join(choices) + '"'
        rule             = DataValidation(type = 'list', formula1 = valid_options, allow_blank = True)
        rule.error       = 'Your entry is not valid'
        rule.errorTitle  = 'Invalid Entry'
        rule.prompt      = 'Please select from list'
        rule.promptTitle = 'Select option'
        sheet.add_data_validation(rule)
        rule.add(cell)
        cell.value       = choices[0]

    def save(self, link:str):
        print(f"saving {link}...")
        self.wb.save(link)

    def isMergeCell(self, i_sheet, i_cell):
        is_merged_cell = False
        for merged_range in i_sheet.merged_cells.ranges:
            if i_cell.coordinate in merged_range:
                is_merged_cell = True
                break
        return is_merged_cell

    def set_value(self, sheet_name:str, cell_addr:str, str_value:str):
        self.wb[sheet_name][cell_addr].value = str_value

    def set_value_by_index(self, irow:int, icol:int, ivalue):
        worksheet.cell(row=irow, column=icol, value=value)

    def get_value(self, sheet_name:str, cell_addr:str):
        return self.wb[sheet_name][cell_addr].value

    def get_cells_from_sheet(self, sheet):
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                cells.append(cell)
        return cells

    def get_column_letter(self, cell):
        return openpyxl.utils.get_column_letter(cell.column)

    def create_sheet(self, sheet_name):
        print(f"Creating sheet: {sheet_name}...")
        new_sheet = self.wb.create_sheet(title = sheet_name)
    
    def create_excel_file(excel_link):
        workbook = openpyxl.Workbook()
        workbook.save(excel_link)

    def remove_sheet(self, sheet_name):
        self.wb.remove(self.get_sheet_by_name(sheet_name))

    def fmt_hyperlink(self, cell, dst_sheet:str, dst_addr:str, message:str):
        #cell.value = f'=HYPERLINK("#{dst_sheet}!{dst_addr}", "{message}")'
        cell.value = f'{message}'
        hyperlink = Hyperlink(ref=f"#{dst_sheet}!{dst_addr}", location=f'{dst_sheet}!{dst_addr}')
        cell.hyperlink = hyperlink
        cell.font  = Font(color='0000FF') #Blue
        #cell.font  = Font(color='0000FF', underline='single') #Blue

    # align  : 'left'    'right'     'center'
    # istyle : 'medium'  'thin'
    def fmt_boder_cell(self, cell, align='center', istyle='medium'):
        cell.border          = Border(left = Side(style=istyle), right = Side(style=istyle), top = Side(style=istyle), bottom = Side(style=istyle))
        if   (align == 'left' ):
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'left',   vertical = 'center')
        elif (align == 'right'):
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'right',  vertical = 'center')
        else:
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    def fmt_alignment_cell(self, cell, align='center'):
        if   (align == 'left' ):
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'left',   vertical = 'center')
        elif (align == 'right'):
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'right',  vertical = 'center')
        else:
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    def fmt_freeze_cell(self, cell):
        sheet = cell.parent
        sheet.freeze_panes = cell.coordinate 

    # input: fmt_table("A1:B12")
    def fmt_table(self, ws, table_range:str):
        # Create a table
        tab = Table(displayName=f"Table{self.table_idx}", ref=table_range)
        # Add a style to the table
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
                                                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        # Add the table to the worksheet
        ws.add_table(tab)
        self.table_idx += 1

    def fmt_table_allcell(self, ws):
        last_row    = ws.max_row
        last_column = ws.max_column
        last_cell   = ws.cell(row   = last_row, column = last_column)
        self.fmt_table(ws, f"A1:{last_cell.coordinate}")

    def get_color(self, color='green'):
        if  (color == 'green'):
            self.font       = Font(color='227447')
            background = 'C6EFCE'
        elif (color == 'yellow'):
            self.font       = Font(color='9C6500')
            background = 'FFEB9C'
        elif (color == 'red'):
            self.font       = Font(color='B71C32')
            background = 'FFC7CE'
        elif (color == 'gray' or color == 'grey'):
            self.font       = Font(color='000000')
            background = 'A5A5A5'
        elif (color == 'blue'):
            self.font  = Font(color='000000')
            background = '95B3D7'
        elif (color == 'header'):
            self.font       = Font(color='000000' )
            background      = '95B3D7'
        self.color_fill = PatternFill(start_color = background, end_color = background, fill_type = 'solid')

    def fmt_cell_color(self, cell, color='green'):
        self.get_color(color)
        cell.font  = self.font
        cell.fill  = self.color_fill

    def fmt_cell_color_bg(self, cell, color='gray'):
#        cell.fill  = PatternFill(start_color = '808080', end_color = '808080', fill_type = 'solid') 
        cell.fill  = PatternFill(start_color = 'CCCCCC', end_color = 'CCCCCC', fill_type = 'solid') 

    def fmt_color_condition(self, cell, keyword:str, color:str):
        sheet = cell.parent
        self.get_color(color)
        rule = FormulaRule(formula=[f'NOT(ISERROR(SEARCH("{keyword}",{cell.coordinate})))'], stopIfTrue=True, fill=self.color_fill, font=self.font)
        sheet.conditional_formatting.add(cell.coordinate, rule)  

    def compare_xlsx(excel_link0, excel_link1):
        all_sheet_cmp = []
        all_sheet     = []

        print(f"\n{excel_link0} is loading ... ")
        workbook  = load_workbook(f"{excel_link0}")
        all_sheet = workbook.get_sheet_names()
        all_sheet.sort()

        try: 
            print(f"\n{excel_link1} is loading ... ")
            workbook_cmp  = load_workbook(f"{excel_link1}")
            all_sheet_cmp = workbook_cmp.get_sheet_names()
            all_sheet_cmp.sort()
        except:
            print(f"{excel_link1} is not existed ... ")

        if (all_sheet != all_sheet_cmp):
            # compare sheet
            print(f"-------------------------------------------")
            print(f"compare number of sheet"                    )
            print(f"-------------------------------------------")
            print(f"    {YELLOW}WARNING: {RESET}{excel_link0} Number of sheet was changed")
            print(f"                        {RESET}{excel_link0} {len(all_sheet)} sheet")
            print(f"                        {RESET}{excel_link1} {len(all_sheet_cmp)} sheet")

        for sheet_name in all_sheet_cmp:
            sheet_is_exist = 0
            print("------------------------------")
            print(sheet_name)
            print("------------------------------")
            sheet_cmp = workbook_cmp[sheet_name] #new
            try:
                sheet     = workbook[sheet_name]
                sheet_is_exist = 1
            except:
                if (sheet_name not in all_sheet):
                    print(f"{sheet_name} was added in {excel_link1}")

            if (sheet_is_exist == 1):
                for row in sheet.iter_rows():
                    for cell in row:
                        cell_addr = cell.coordinate
                        if (str(sheet[cell_addr].value) != str(sheet_cmp[cell_addr].value)):
                            # compare data
                            print(f"    {excel_link0}: {sheet_name}.{cell_addr} was changed"  )
                            print(f"        {cell.coordinate}: {sheet[cell_addr].value}"     )
                            print(f"        {cell.coordinate}: {sheet_cmp[cell_addr].value}" )

    def remove_trash_cell(self):
        print("Remove trash cell")
        for sheet_name in self.sheet_names:
            sheet = self.wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell != None:
                        all_cell_as_space = 1
                        for ch in str(cell.value):
                            if ch != " ":
                                all_cell_as_space = 0
                        if (all_cell_as_space == 1):
                            print(f"remove {sheet_name}.{cell.coordinate}")
                            cell.value = None
        self.save(self.link)

    def convert_xls_to_xlsx(input_file):
        if (tian.get_file_type(input_file) == ".xlsx"):
            return 0
        elif (tian.get_file_type(input_file) != ".xls"):
            print("Error: Only be able to convert .xls -> .xlsx ")
            print(f"Filename: {input_file}")
            exit()

        output_file = input_file.replace(".xls", ".xlsx")
        print(f"convert {input_file} --> {output_file}")
        try:
            # Read the xls file using xlrd
            xls_workbook    = xlrd.open_workbook(input_file, formatting_info = True)
            xls_sheet_names = xls_workbook.sheet_names()

            # Create a new xlsx workbook
            xlsx_workbook    = openpyxl.Workbook()
            xlsx_sheet_names = xlsx_workbook.sheetnames

            # Copy data from xls to xlsx
            for sheet_name in xls_sheet_names:
                xls_sheet  = xls_workbook.sheet_by_name(sheet_name)
                xlsx_sheet = xlsx_workbook.create_sheet(title       = sheet_name)

                for row_index in range(xls_sheet.nrows):
                    row_data = xls_sheet.row_values(row_index)

                    for col_index, cell_value in enumerate(row_data):
                        if (cell_value == ""):
                            continue
                        cell                 = xlsx_sheet.cell(row = row_index + 1, column = col_index + 1)
                        cell.value           = cell_value
                        cell.border          = Border(left = Side(style='medium'), right = Side(style='medium'), top = Side(style='medium'), bottom = Side(style='medium'))
                        cell.alignment       = openpyxl.styles.Alignment(horizontal = 'center', vertical    = 'center')

            # Copy merge addr
            for sheet_name in xls_sheet_names:
                xls_sheet  = xls_workbook.sheet_by_name(sheet_name)
                xlsx_sheet = xlsx_workbook[sheet_name]

                merged_cells = xls_sheet.merged_cells
                for merged_cell in merged_cells:
                    start_row, end_row, start_col, end_col = merged_cell
                    # Construct the address of the merged cell range
                    cell_range = xlrd.cellname(start_row, start_col) + ':' + xlrd.cellname(end_row - 1, end_col - 1)
                    xlsx_sheet.merge_cells(cell_range)

            xlsx_workbook.remove(xlsx_workbook['Sheet'])

            # Save the new xlsx file
            xlsx_workbook.save(output_file)

            print_banner(f"Copying successful. File saved as {output_file}")
        except Exception as e:
            print(f"Error during copying: {e}")


#--------------------------------------------------------------------+
# Note                                                               |
#--------------------------------------------------------------------+
# group
"""
for row in range(1, 9):
    new_sheet.row_dimensions[row].outline_level = 1

for col in col_symbol:
    new_sheet.column_dimensions[col].outline_level = 1
"""

