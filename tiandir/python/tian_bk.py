#!/usr/bin/python3
#--------------------------------------------------------------------
#  File name     : tian.py
#  Function      : <empty>
#  Author        : Tian
#  Email         : tientq@coasia.com
#  Created time  : 2024-02-16 15:26:37
#--------------------------------------------------------------------
import re, os, sys, readline
sys.path.append("/mnt/users/tientq/.local/lib/python3/site-packages") 
sys.path.append("/mnt/users/tientq/python_dir") 
import openpyxl
from openpyxl import load_workbook
import xlrd
import subprocess

#-------------------------------------------------------------------|
# Print color                                                       |
#-------------------------------------------------------------------|
BLACK   = '\033[30;1m'
RED     = '\033[31;1m'
GREEN   = '\033[32;1m'
YELLOW  = '\033[33;1m'
BLUE    = '\033[34;1m'
MAGENTA = '\033[35;1m'
CYAN    = '\033[36;1m'
WHITE   = '\033[37;1m'
RESET   = '\033[0m'

#-------------------------------------------------------------------#
# Function                                                          #
#-------------------------------------------------------------------#
def run(command:str):
    print(f">>> {command}")
    os.system(command)

run_process = 0
def run_with_output(command:str):
    global run_process
    run_process += 1
    os.system(f"{command} > tmp_output_{run_process}")
    with open(f'tmp_output_{run_process}', 'r') as f:
        rdata = f.read()
    os.system(f"rm -rf tmp_output_{run_process}")

    if type(rdata) == int:
        return rdata
    if type(rdata) == str:
        rdata = rdata.split('\n')
        # return string
        if (len(rdata) == 2) and (rdata[1] == ''):
            return rdata[0]
        # return list
        else:
            return rdata

def read_file(link_file):
    return file.read(link_file)

def write_file(link_file, wdata):
    file.write(link_file, wdata)

def get_folder_names(path):
    folder_names = []
    for entry in os.scandir(path):
        if entry.is_dir():
            folder_names.append(entry.name)
    return folder_names



# ouput: hex string
#-------------------------------------
def b2h(bin_string):
    decimal_str = int(bin_string, 2)
    hexa = hex(decimal_str)
    return hexa.replace("0x", "")

def h2d(hex_string):
    int(hex_string, 16)

# output: string
# -------------------------------------
def current_folder():
    return os.path.basename(os.getcwd())

def sub_in_list(list_name, old, new):
    out = []
    for line in list_name:
        line = re.sub(old, new, line)
        out.append(line)
    return out

def print_banner(string_value):
    string_value = "#  " + string_value
    print(f"{BLUE}#" + len(string_value)*"-" + "-#")
    print(string_value + f"{BLUE}  #")
    print(f"{BLUE}#" + len(string_value)*"-" + "-#" + RESET)


def get_env_var(env_var=""):
    return os.environ.get(env_var)


def is_file(path_str=""):
    if ((os.path.exists(path_str)) and (os.path.isfile(path_str)) == True):
        return True
    else:
        return False

def is_dir(path_str=""):
    if ((os.path.exists(path_str)) and (os.path.isdir(path_str)) == True):
        return True
    else:
        return False

#
# Function: print_class_attr
#
def print_class_attr(class_object):
    import inspect
    all_attrs = dir(class_object)
    print("Methods:")
    for attr in all_attrs:
        if ("__" in attr):
            continue
        if callable(getattr(class_object, attr)):
            print("\n   "+ attr + "()")
            parameters = inspect.signature(getattr(class_object, attr)).parameters
            for param in parameters:
                print("   |--- " + "%-30s"%(param) + str(parameters[param].annotation))
    print("Properties:")
    for attr in all_attrs:
        if ("__" in attr):
            continue
        if callable(getattr(class_object, attr)):
            continue
        print("   "+ attr)

def print_list(ilist:list):
    for item in ilist:
        print(item)

# ------------------------------------------------------------------------------|
# Class: file                                                                   |
# It is a static class (dont create this class)                                 | 
# ------------------------------------------------------------------------------|
class file:
    def read(link):
        output = []
        print(f"reading from {link} ...")
        with open(link, 'r') as f:
            lines = f.readlines()

        for line in lines:
            output.append(line)
        return output

    def write(link, wdata):
        w = open(link, "w")
        print(f"writing to {link} ...")
        if   (type(wdata) == str):
            w.write(wdata)
        elif (type(wdata) == list):
            for line in wdata:
                w.write(line)

    def append(link, wdata):
        w = open(link, "a")
        print(f"writing to {link} ...")
        if   (type(wdata) == str):
            w.write(wdata)
        elif (type(wdata) == list):
            for line in wdata:
                w.write(line)

    # must use [a-z]*, [0-9]* to present regular expression
    def replace(link, old:str, new:str):
        print("{}   :   {}   -->   {}".format(link, old, new))
        old = old.replace("/", "\/"); old = old.replace(".", "\.")
        new = new.replace("/", "\/"); new = new.replace(".", "\.")
        os.system("sed -i 's/{}/{}/g' {}".format(old, new, link_file))

    def make_uniq(link):
        os.system(f"sort {link} | uniq > {link}_uniq")

    def remove_line(link, keyword):
        fileout = []
        filein = read(link)
        for line in filein:
            if (key_remove in line):
                continue
            fileout.append(line)
        write(link, fileout)

    def execute(link):
        realpath = run_with_output(f"realpath {link}")
        os.system(f"chmod 755 {realpath}")
        os.system(f"{realpath}")

# --------------------------------------------------------------------|
#  Class                                                              |
# --------------------------------------------------------------------|
from openpyxl                          import formatting, styles
from openpyxl.formatting.rule          import FormulaRule, CellIsRule
from openpyxl.styles                   import PatternFill, Font, NamedStyle
from openpyxl.styles.borders           import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers           import FORMAT_PERCENTAGE_00
class excel:
    # ----------------------------------------------------------------|
    # Properties                                                      |
    # ----------------------------------------------------------------|
    font        = Font(color              = '227447')
    color_fill  = PatternFill(start_color = 'C6EFCE', end_color = 'C6EFCE', fill_type = 'solid')
    link        = ""
    sheet_names = []
    wb          = None

    # ----------------------------------------------------------------|
    # Methods                                                         |
    # ----------------------------------------------------------------|
    def __init__(self, excel_link):
        #xls --> xlsx
        if ("xlsx" not in excel_link):
            print(f"Info: Cant not open xls format directly")
            print(f"      Automatic converting to xlsx format")
            print(f"      Please handles new file: {excel_link}x")
            self.convert_xls_to_xlsx(excel_link, f"{excel_link}x")
            excel_link = excel_link + "x"

        self.link = excel_link
        try:
            self.wb          = load_workbook(f"{excel_link}")
            self.sheet_names = self.wb.get_sheet_names()
            print_banner(f"{BLUE}{excel_link} is loading ...")
        except:
            print(f"open {excel_link} error!")
            os.system(f"echo {excel_link} >> file_error.log")

    def get_sheet_by_name(self, sheet_name:str):
        try:
            return self.wb[sheet_name]
        except:
            print(f"{self.link}.{sheet_name} is not exist")

    def resize_col(self, sheet_name:str, col_str:str, width:int):
        self.wb[sheet_name].column_dimensions[col_str].width = width

    def resize_row(self, sheet_name:str, row_int:int, height:int):
        self.wb[sheet_name].row_dimensions[row_int].height = height

    def fmt_wrap_text(self, cell):
        cell.alignment = openpyxl.styles.Alignment(wrapText=True)
    
    def fmt_percent_style(self, cell):
        sheet              = cell.parent
        cell.number_format = '0.00%'
        self.get_color('red')
        rule_red           = CellIsRule(operator='between', formula=['0'  , '0.5'], stopIfTrue=True, fill=self.color_fill, font=self.font)
        self.get_color('yellow')
        rule_yellow        = CellIsRule(operator='between', formula=['0.5', '0.8'], stopIfTrue=True, fill=self.color_fill, font=self.font)
        self.get_color('green')
        rule_green         = CellIsRule(operator='between', formula=['0.8', '1'  ], stopIfTrue=True, fill=self.color_fill, font=self.font)
        sheet.conditional_formatting.add(cell.coordinate, rule_red)
        sheet.conditional_formatting.add(cell.coordinate, rule_yellow)
        sheet.conditional_formatting.add(cell.coordinate, rule_green)
    
    def drop_down_list(self, cell, choices:list):
        sheet            = cell.parent
        valid_options    = '"' + ','.join(choices) + '"'
        rule             = DataValidation(type = 'list', formula1 = valid_options, allow_blank = True)
        rule.error       = 'Your entry is not valid'
        rule.errorTitle  = 'Invalid Entry'
        rule.prompt      = 'Please select from list'
        rule.promptTitle = 'Select option'
        sheet.add_data_validation(rule)
        rule.add(cell)
        cell.value       = choices[0]

    def save(self, f_name:str):
        print(f"saving {self.link}")
        self.wb.save(f_name)

    def isMergeCell(self, i_sheet:openpyxl.worksheet.worksheet.Worksheet, i_cell:openpyxl.cell.cell.Cell):
        is_merged_cell = False
        for merged_range in i_sheet.merged_cells.ranges:
            if i_cell.coordinate in merged_range:
                is_merged_cell = True
                break
        return is_merged_cell

    def set_value(self, sheet_name:str, cell_addr:str, str_value:str):
        self.wb[sheet_name][cell_addr].value = str_value

    def get_value(self, sheet_name:str, cell_addr:str):
        return self.wb[sheet_name][cell_addr].value

    def get_cells_from_sheet(self, sheet):
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                cells.append(cell)
        return cells

    def get_column_letter(self, cell):
        return openpyxl.utils.get_column_letter(cell.column)

    def create_sheet(self, sheet_name):
        new_sheet = self.wb.create_sheet(title = sheet_name)
    
    def create_excel_file(self, excel_link):
        workbook = openpyxl.Workbook()
        workbook.save(excel_link)

    def remove_sheet(self, sheet_name):
        self.wb.remove(self.get_sheet_by_name(sheet_name))

    def fmt_hyperlink(self, cell, dst_sheet:str, dst_addr:str, message:str, filetype='microsoft'):
        if (filetype == 'microsoft'):
            cell.value = f'=HYPERLINK("#{dst_sheet}!{dst_addr}", "{message}")'
            cell.font  = Font(color='0000FF', underline='single') #Blue
        elif (filetype == 'googlesheet'):
            print("fmt_hyperlink for googlesheet")
            cell.value = f'=HYPERLINK("#gid=" & sheetid("{dst_sheet}"); "{message}")'
            cell.font  = Font(color='0000FF', underline='single') #Blue

    def fmt_boder_cell(self, cell, align='center'):
        cell.border          = Border(left = Side(style='medium'), right = Side(style='medium'), top = Side(style='medium'), bottom = Side(style='medium'))
        if   (align == 'left' ):
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'left',   vertical = 'center')
        elif (align == 'right'):
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'right',  vertical = 'center')
        else:
            cell.alignment       = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    def get_color(self, color='green'):
        if  (color == 'green'):
            self.font       = Font(color='227447')
            background = 'C6EFCE'
        elif (color == 'yellow'):
            self.font       = Font(color='9C6500')
            background = 'FFEB9C'
        elif (color == 'red'):
            self.font       = Font(color='B71C32')
            background = 'FFC7CE'
        elif (color == 'gray' or color == 'grey'):
            self.font       = Font(color='000000')
            background = 'A5A5A5'
        elif (color == 'header'):
            self.font       = Font(color='000000' )
            background      = '95B3D7'
        self.color_fill = PatternFill(start_color = background, end_color = background, fill_type = 'solid')

    def fmt_cell_color(self, cell, color='green'):
        self.get_color(color)
        cell.font  = self.font
        cell.fill  = self.color_fill

    def fmt_color_condition(self, cell, keyword:str, color:str):
        sheet = cell.parent
        self.get_color(color)
        rule = FormulaRule(formula=[f'NOT(ISERROR(SEARCH("{keyword}",{cell.coordinate})))'], stopIfTrue=True, fill=self.color_fill, font=self.font)
        sheet.conditional_formatting.add(cell.coordinate, rule)  

    def compare_xlsx(excel_link0, excel_link1):
        all_sheet_cmp = []
        all_sheet     = []

        print(f"\n{excel_link0} is loading ... ")
        workbook  = load_workbook(f"{excel_link0}")
        all_sheet = workbook.get_sheet_names()
        all_sheet.sort()

        try: 
            print(f"\n{excel_link1} is loading ... ")
            workbook_cmp  = load_workbook(f"{excel_link1}")
            all_sheet_cmp = workbook_cmp.get_sheet_names()
            all_sheet_cmp.sort()
        except:
            print(f"{excel_link1} is not existed ... ")

        if (all_sheet != all_sheet_cmp):
            # compare sheet
            print(f"-------------------------------------------")
            print(f"compare number of sheet"                    )
            print(f"-------------------------------------------")
            print(f"    {YELLOW}WARNING: {RESET}{excel_link0} Number of sheet was changed")
            print(f"                        {RESET}{excel_link0} {len(all_sheet)} sheet")
            print(f"                        {RESET}{excel_link1} {len(all_sheet_cmp)} sheet")

        for sheet_name in all_sheet_cmp:
            sheet_is_exist = 0
            print("------------------------------")
            print(sheet_name)
            print("------------------------------")
            sheet_cmp = workbook_cmp[sheet_name] #new
            try:
                sheet     = workbook[sheet_name]
                sheet_is_exist = 1
            except:
                if (sheet_name not in all_sheet):
                    print(f"{sheet_name} was added in {excel_link1}")

            if (sheet_is_exist == 1):
                for row in sheet.iter_rows():
                    for cell in row:
                        cell_addr = cell.coordinate
                        if (str(sheet[cell_addr].value) != str(sheet_cmp[cell_addr].value)):
                            # compare data
                            print(f"    {excel_link0}: {sheet_name}.{cell_addr} was changed"  )
                            print(f"        {cell.coordinate}: {sheet[cell_addr].value}"     )
                            print(f"        {cell.coordinate}: {sheet_cmp[cell_addr].value}" )

    def remove_trash_cell(self):
        print("Remove trash cell")
        for sheet_name in self.sheet_names:
            sheet = self.wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell != None:
                        all_cell_as_space = 1
                        for ch in str(cell.value):
                            if ch != " ":
                                all_cell_as_space = 0
                        if (all_cell_as_space == 1):
                            print(f"remove {sheet_name}.{cell.coordinate}")
                            cell.value = None
        self.save(self.link)

    def convert_xls_to_xlsx(self, input_file, output_file):
        print(f"convert {input_file} --> {output_file}")
        try:
            # Read the xls file using xlrd
            xls_workbook    = xlrd.open_workbook(input_file, formatting_info = True)
            xls_sheet_names = xls_workbook.sheet_names()

            # Create a new xlsx workbook
            xlsx_workbook    = openpyxl.Workbook()
            xlsx_sheet_names = xlsx_workbook.sheetnames

            # Copy data from xls to xlsx
            for sheet_name in xls_sheet_names:
                xls_sheet  = xls_workbook.sheet_by_name(sheet_name)
                xlsx_sheet = xlsx_workbook.create_sheet(title       = sheet_name)

                for row_index in range(xls_sheet.nrows):
                    row_data = xls_sheet.row_values(row_index)

                    for col_index, cell_value in enumerate(row_data):
                        if (cell_value == ""):
                            continue
                        cell                 = xlsx_sheet.cell(row = row_index + 1, column = col_index + 1)
                        cell.value           = cell_value
                        cell.border          = Border(left = Side(style='medium'), right = Side(style='medium'), top = Side(style='medium'), bottom = Side(style='medium'))
                        cell.alignment       = openpyxl.styles.Alignment(horizontal = 'center', vertical    = 'center')

            # Copy merge addr
            for sheet_name in xls_sheet_names:
                xls_sheet  = xls_workbook.sheet_by_name(sheet_name)
                xlsx_sheet = xlsx_workbook[sheet_name]

                merged_cells = xls_sheet.merged_cells
                for merged_cell in merged_cells:
                    start_row, end_row, start_col, end_col = merged_cell
                    # Construct the address of the merged cell range
                    cell_range = xlrd.cellname(start_row, start_col) + ':' + xlrd.cellname(end_row - 1, end_col - 1)
                    xlsx_sheet.merge_cells(cell_range)

            xlsx_workbook.remove(xlsx_workbook['Sheet'])

            # Save the new xlsx file
            xlsx_workbook.save(output_file)

            print_banner(f"Copying successful. File saved as {output_file}")
        except Exception as e:
            print(f"Error during copying: {e}")

# ------------------------------------------------------------------------------|
# Class: argument_vector is a static class                                      |
# ------------------------------------------------------------------------------|
class argument_vector:
    def len():
        return len(sys.argv)

    def has_argument(arg_mess:str):
        if arg_mess in sys.argv:
            print(f"has argument: {arg_mess}")
            return True
        else:
            print(f"Dont has argument: {arg_mess}")
            return False

    def get_argument_by_prefix(prefix_arg_mess:str):
        if prefix_arg_mess in sys.argv:
            result = sys.argv[sys.argv.index(prefix_arg_mess)+1]
            print(f"get argument {prefix_arg_mess}: {result}")
        else:
            print(f"Dont has argument: {prefix_arg_mess}")

    def print():
        i = 0
        for item in sys.argv:
            print(f"Argument[{i}] = {sys.argv[i]}")
            i += 1

# ------------------------------------------------------------------------------|
# Class: time is a static class                                                 |
# ------------------------------------------------------------------------------|
from datetime import datetime 
class time:
    def full():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    def hour():
        return datetime.now().strftime("%H:%M:%S")
    def day():
        return datetime.now().strftime("%Y-%m-%d")


#------------------------------------------------------#
# Note
#------------------------------------------------------#
#    with open('done.sv', 'w') as f:
#        f.write(<file_link>)
#    with open('done.sv', 'r') as f:
#        rdata = f.read(<file_link>)
#   os.getcwd() - get pwd
#   sys.argv    - list argv
#   all_sheet = workbook.get_sheet_names()
